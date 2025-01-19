from alunos import cadastrar_aluno, listar_alunos, adicionar_notas, listar_notas
from professores import cadastrar_professor, listar_professores
from turmas import criar_turma, listar_turmas
from fpdf import FPDF
import openpyxl
import os

alunos = [] 
professores = [] 
turmas = [] 

def salvar_em_pdf(alunos, professores, turmas):
  
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

  
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(200, 10, txt="Relatório de Gestão Escolar", ln=True, align="C")
    pdf.ln(10)

  
    pdf.set_font("Arial", style="B", size=12)
    pdf.cell(200, 10, txt="--- Alunos ---", ln=True)
    pdf.set_font("Arial", size=11)
    if alunos:
        for aluno in alunos:
            disciplinas = ', '.join(aluno['disciplinas'])
            pdf.cell(200, 10, txt=f"Nome: {aluno['nome']}, Idade: {aluno['idade']}, Ano Escolar: {aluno['ano_escolar']}", ln=True)
            pdf.cell(200, 10, txt=f"  Disciplinas: {disciplinas}", ln=True)
            if aluno['notas']:
                for disciplina, info in aluno['notas'].items():
                    notas = ', '.join(map(str, info['notas']))
                    pdf.cell(200, 10, txt=f"    {disciplina} - Notas: {notas} - Status: {info['status']}", ln=True)
            pdf.ln(5)
    else:
        pdf.cell(200, 10, txt="Nenhum aluno cadastrado.", ln=True)

 
    pdf.set_font("Arial", style="B", size=12)
    pdf.cell(200, 10, txt="--- Professores ---", ln=True)
    pdf.set_font("Arial", size=11)
    if professores:
        for professor in professores:
            especialidades = ', '.join(professor['especialidade'])
            horarios = ', '.join(professor['horarios'])
            pdf.cell(200, 10, txt=f"Nome: {professor['nome']}, Especialidade: {especialidades}", ln=True)
            pdf.cell(200, 10, txt=f"  Horários Disponíveis: {horarios}", ln=True)
            pdf.ln(5)
    else:
        pdf.cell(200, 10, txt="Nenhum professor cadastrado.", ln=True)

  
    pdf.set_font("Arial", style="B", size=12)
    pdf.cell(200, 10, txt="--- Turmas ---", ln=True)
    pdf.set_font("Arial", size=11)
    if turmas:
        for turma in turmas:
            pdf.cell(200, 10, txt=f"Nome: {turma['nome']}, Professor: {turma['professor']['nome']}", ln=True)
            if 'alunos' in turma and turma['alunos']:
                pdf.cell(200, 10, txt="  Alunos:", ln=True)
                for aluno in turma['alunos']:
                    pdf.cell(200, 10, txt=f"    - {aluno['nome']} ({aluno['ano_escolar']})", ln=True)
            else:
                pdf.cell(200, 10, txt="  Nenhum aluno associado à turma.", ln=True)

   
    nome_arquivo = "Relatorio_Gestao_Escolar.pdf"
    pdf.output(nome_arquivo)
    print(f"Relatório salvo em PDF: {nome_arquivo}")

    
    os.startfile(nome_arquivo)
    
def salvar_em_excel(alunos, professores, turmas):
 
    wb = openpyxl.Workbook()

   
    ws_alunos = wb.active
    ws_alunos.title = "Alunos"
    ws_alunos.append(["Nome", "Idade", "Ano Escolar", "Disciplinas", "Notas e Status"])
    for aluno in alunos:
        disciplinas = ', '.join(aluno['disciplinas'])
        notas_info = ""
        if aluno['notas']:
            for disciplina, info in aluno['notas'].items():
                notas = ', '.join(map(str, info['notas']))
                notas_info += f"{disciplina}: {notas} ({info['status']})\n"
        ws_alunos.append([aluno['nome'], aluno['idade'], aluno['ano_escolar'], disciplinas, notas_info.strip()])

   
    ws_professores = wb.create_sheet("Professores")
    ws_professores.append(["Nome", "Especialidade", "Horários Disponíveis"])
    for professor in professores:
        especialidades = ', '.join(professor['especialidade'])
        horarios = ', '.join(professor['horarios'])
        ws_professores.append([professor['nome'], especialidades, horarios])

  
    ws_turmas = wb.create_sheet("Turmas")
    ws_turmas.append(["Nome da Turma", "Professor Responsável", "Alunos"])
    for turma in turmas:
        alunos_info = ""
        if 'alunos' in turma and turma['alunos']:
            for aluno in turma['alunos']:
                alunos_info += f"{aluno['nome']} ({aluno['ano_escolar']}), "
        ws_turmas.append([turma['nome'], turma['professor']['nome'], alunos_info.strip(", ")])

    nome_arquivo = "Relatorio_Gestao_Escolar.xlsx"
    wb.save(nome_arquivo)
    print(f"Relatório salvo em Excel: {nome_arquivo}")

    os.startfile(nome_arquivo)